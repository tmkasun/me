from base64 import b64encode  # To convert username:password to base64 encoded string
import requests  # Python simple HTTP library
import openpyxl  # Python Excel file reader library
from configs import conf  # Loads configuration details of the tool


def main():
    wb = openpyxl.load_workbook(conf['excel_file'])
    name = wb.get_sheet_names()[0]
    sheet = wb.get_sheet_by_name(name)
    for row in range(2, sheet.max_row):  # from 2 (assuming first row contains headers) to end of the row
        condition = sheet.cell(row=row, column=1).value
        color = sheet.cell(row=row, column=2).value
        year = sheet.cell(row=row, column=3).value
        name = sheet.cell(row=row, column=4).value
        description = sheet.cell(row=row, column=5).value
        body = sheet.cell(row=row, column=6).value
        notes = sheet.cell(row=row, column=7).value
        temp = populate_template(condition, color, year, name, description, body, notes)
        asset_name = name.replace(" ", "_")
        create_asset(temp, asset_name)


def create_asset(metadata, file_name):
    credentials = '{username}:{password}'.format(username=conf['username'], password=conf['password'])
    auth_value = b64encode(credentials.encode())
    headers = {
        'Authorization': '{type} {auth_value}'.format(type="Basic", auth_value=auth_value.decode()),
        'Content-Type': conf['media-type'],
        'Cache-Control': 'no-cache'
    }
    request_url = conf['artifact_endpoint'] + conf['reg_path'] + file_name
    response = requests.put(request_url, data=metadata, headers=headers, verify=False)
    if not response.ok:
        print("Error while creating the asset {}".format(file_name))
    else:
        print("Response Content = {}\nStatus Code = {}".format(response.content, response.status_code))


def populate_template(condition, color, year, name, description, body, notes):
    template = """<?xml version="1.0"?>
    <metadata xmlns="http://www.wso2.org/governance/metadata">
      <overview>
        <condition>{}</condition>
        <color>{}</color>
        <year>{}</year>
        <name>{}</name>
        <description>{}</description>
        <body>{}</body>
      </overview>
      <notes>
        <notes>{}</notes>
      </notes>
    </metadata>""".format(condition, color, year, name, description, body, notes)
    return template


if __name__ == '__main__':
    main()
